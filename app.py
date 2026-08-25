import streamlit as st
import pandas as pd
import os
import glob
import sys
import chardet
import openpyxl
from openpyxl.utils import cell
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import shutil
import io
import numpy as np

# ======================================================
# 💡 設定: ファイル名
# ======================================================
EXCEL_TEMPLATE_FILENAME = 電力報告テンプレート.xlsx'


# --- CSV読み込み関数 (エンコーディング自動検出) ---
@st.cache_data
def detect_and_read_csv(uploaded_file):
    """アップロードされたファイルの内容を読み込み、エンコーディングを自動検出してDataFrameを返す"""
    uploaded_file.seek(0)
    raw_data = uploaded_file.read()
    
    detected_encoding = chardet.detect(raw_data)['encoding']
    encodings_to_try = ['cp932', 'shift_jis', 'utf-8']
    
    if detected_encoding and detected_encoding.lower() not in encodings_to_try:
        encodings_to_try.append(detected_encoding.lower())

    for encoding in encodings_to_try:
        try:
            df = pd.read_csv(io.BytesIO(raw_data), header=1, encoding=encoding) 
            if '年' in df.columns:
                return df
            else:
                continue
        except Exception:
            continue
            
    raise Exception(f"ファイル '{uploaded_file.name}' は、一般的な日本語エンコーディングで読み込めませんでした。")


# --- Excelレポート書き込み関数 ---
def write_excel_reports(excel_file_path, df_before, df_after, start_before, end_before, start_after, end_after, operating_hours, store_name):
    
    SHEET1_NAME = 'Sheet1'
    SUMMARY_SHEET_NAME = 'まとめ'
    
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        st.error(f"エラー: Excelテンプレートが見つかりません。")
        return False

    # --- 共通計算 ---
    days_before = (end_before - start_before).days + 1
    days_after = (end_after - start_after).days + 1
    
    # 1日あたりの平均合計kWh (期間内の合計kWh ÷ 日数)
    total_kWh_before = df_before['合計kWh'].sum()
    total_kWh_after = df_after['合計kWh'].sum()
    
    avg_daily_total_before = total_kWh_before / days_before if days_before > 0 and not np.isnan(total_kWh_before) else 0
    avg_daily_total_after = total_kWh_after / days_after if days_after > 0 and not np.isnan(total_kWh_after) else 0
    
    # --- 1. Sheet1: 24時間別平均の書き込み (C36～D59) と合計値 (C33, D33) ---
    if SHEET1_NAME not in workbook.sheetnames:
        workbook.create_sheet(SHEET1_NAME) 
        
    ws_sheet1 = workbook[SHEET1_NAME]
    
    ws_sheet1['C33'] = float(avg_daily_total_before)
    ws_sheet1['D33'] = float(avg_daily_total_after)
    
    # 24時間別平均（各時間帯での平均kWh/h）
    metrics_before = df_before.groupby('時')['合計kWh'].mean() if not df_before.empty else None
    metrics_after = df_after.groupby('時')['合計kWh'].mean() if not df_after.empty else None

    current_row = 36
    for start_hour in range(0, 24):
        end_hour = (start_hour + 1) % 24
        time_range = f"{start_hour:02d}:00～{end_hour:02d}:00"

        ws_sheet1.cell(row=current_row, column=1, value=f"{start_hour:02d}:00") 
        ws_sheet1.cell(row=current_row, column=2, value=time_range) 
        
        # C列 (施工前 平均)
        value_before = 0.0
        if metrics_before is not None and start_hour in metrics_before.index:
            mean_val = metrics_before.loc[start_hour]
            value_before = float(mean_val) if not np.isnan(mean_val) else 0.0
        ws_sheet1.cell(row=current_row, column=3, value=value_before)
            
        # D列 (施工後 平均)
        value_after = 0.0
        if metrics_after is not None and start_hour in metrics_after.index:
            mean_val = metrics_after.loc[start_hour]
            value_after = float(mean_val) if not np.isnan(mean_val) else 0.0
        ws_sheet1.cell(row=current_row, column=4, value=value_after)
            
        current_row += 1
    
    ws_sheet1['C35'] = '施工前 平均kWh/h'
    ws_sheet1['D35'] = '施工後 平均kWh/h'
    ws_sheet1['A35'] = '時間帯'

    # --- 2. まとめシートの書き込み ---
    if SUMMARY_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(SUMMARY_SHEET_NAME)
        
    ws_summary = workbook[SUMMARY_SHEET_NAME]

    format_date = lambda d: f"{d.year}/{d.month}/{d.day}"

    start_b_str = format_date(start_before)
    end_b_str = format_date(end_before)
    before_str = f"施工前：{start_b_str}～{end_b_str}（{days_before}日間）"
    
    start_a_str = format_date(start_after)
    end_a_str = format_date(end_after)
    after_str = f"施工後(調光後)：{start_a_str}～{end_a_str}（{days_after}日間）"

    ws_summary['H6'] = before_str
    ws_summary['H7'] = after_str
    ws_summary['H8'] = operating_hours
    ws_summary['B1'] = f"{store_name}の使用電力比較報告書"
    
    ws_summary['B7'] = float(avg_daily_total_before)
    ws_summary['B8'] = float(avg_daily_total_after)
    
    workbook.save(excel_file_path)
    return True


# --- Streamlitメインアプリケーション ---
def main_streamlit_app():
    st.set_page_config(layout="wide", page_title="電力データ報告書作成アプリ")
    st.title("💡 電力データ自動処理アプリ")
    st.markdown("### Step 1: ファイルのアップロード")
    
    # --- 1. CSVファイルのアップロード ---
    uploaded_csvs = st.file_uploader(
        "📈 CSVデータ (複数可) をアップロードしてください",
        type=['csv'],
        accept_multiple_files=True
    )
    
    if uploaded_csvs:
        st.success(f"CSVファイル {len(uploaded_csvs)}個 が準備できました。")
        st.markdown("---")
        st.markdown("### Step 2: 期間と情報の入力")
    else:
        st.warning("処理を開始するには、CSVデータをアップロードしてください。")
        return

    # --- 2. ユーザー入力ウィジェット ---
    today = datetime.date.today()
    
    col_date1, col_date2 = st.columns(2)
    
    with col_date1:
        st.subheader("🗓️ 施工前 測定期間")
        start_before = st.date_input("開始日", today - datetime.timedelta(days=30), key="start_b")
        end_before = st.date_input("終了日", today - datetime.timedelta(days=23), key="end_b")
        
    with col_date2:
        st.subheader("📅 施工後 測定期間")
        start_after = st.date_input("開始日", today - datetime.timedelta(days=14), key="start_a")
        end_after = st.date_input("終了日", today - datetime.timedelta(days=7), key="end_a")

    col_info1, col_info2 = st.columns(2)
    with col_info1:
        operating_hours = st.text_input("営業時間", value="08:00-22:00", help="まとめシートH8に反映")
    with col_info2:
        store_name = st.text_input("店舗名", value="大倉山店", help="報告書名とまとめシートB1に反映")
        
    st.markdown("---")
    
    # --- 3. 実行ボタン ---
    if st.button("🚀 データ処理を実行し、報告書をダウンロード"):
        if start_before >= end_before or start_after >= end_after:
            st.error("🚨 期間の設定が不正です。開始日は終了日よりも前に設定してください。")
            return

        try:
            temp_dir = "temp_data"
            os.makedirs(temp_dir, exist_ok=True)
            
            if not os.path.exists(EXCEL_TEMPLATE_FILENAME):
                st.error(f"🚨 致命的なエラー: Excelテンプレートファイル '{EXCEL_TEMPLATE_FILENAME}' が実行環境から見つかりません。")
                return

            temp_excel_path = os.path.join(temp_dir, EXCEL_TEMPLATE_FILENAME)
            shutil.copy(EXCEL_TEMPLATE_FILENAME, temp_excel_path)
            
            # --- b) データ処理 ---
            file_data_list = []
            
            for csv_file in uploaded_csvs:
                df = detect_and_read_csv(csv_file)
                
                df['年'] = pd.to_numeric(df['年'], errors='coerce').astype('Int64')
                df['月'] = pd.to_numeric(df['月'], errors='coerce').astype('Int64')
                df['日'] = pd.to_numeric(df['日'], errors='coerce').astype('Int64')
                df['時'] = pd.to_numeric(df['時'], errors='coerce').astype('Int64')
                
                df.dropna(subset=['年', '月', '日', '時'], inplace=True)
                
                # 同一ファイル内での同じ日時の重複行除去
                df.drop_duplicates(subset=['年', '月', '日', '時'], keep='first', inplace=True)
                
                # E列以降（回路列）の加算
                datetime_cols = ['年', '月', '日', '時']
                consumption_cols = [col for col in df.columns if col not in datetime_cols and not col.startswith('Unnamed:')]
                
                for col in consumption_cols:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                
                # 1ファイル（1フロア）内の1時間の合計kWh
                df['ファイル内kWh'] = df[consumption_cols].sum(axis=1)
                
                file_data_list.append(df[['年', '月', '日', '時', 'ファイル内kWh']])
            
            # 全ファイルのデータを縦に結合
            df_all = pd.concat(file_data_list, ignore_index=True)
            
            # 💡【重要】各ファイル（1階、2階など）の同じ日時のデータを足し合わせる
            df_combined = df_all.groupby(['年', '月', '日', '時'], as_index=False)['ファイル内kWh'].sum()
            df_combined.rename(columns={'ファイル内kWh': '合計kWh'}, inplace=True)
            
            # 「時」が24（1〜24表記）の場合、0〜23に変換
            if not df_combined.empty and df_combined['時'].max() > 23:
                df_combined['時'] = df_combined['時'] - 1
                st.info("💡 CSVの「時」カラムが1-24形式だったため、0-23形式に標準化しました。")
            
            df_combined['日付'] = pd.to_datetime(
                df_combined['年'].astype(str) + '-' + df_combined['月'].astype('str') + '-' + df_combined['日'].astype('str'), 
                format='%Y-%m-%d', errors='coerce'
            ).dt.date
            df_combined.dropna(subset=['日付'], inplace=True)

            # --- c) 期間分割 ---
            start_b = start_before
            end_b = end_before
            start_a = start_after
            end_a = end_after

            df_before = df_combined[(df_combined['日付'] >= start_b) & (df_combined['日付'] <= end_b)].copy()
            df_after = df_combined[(df_combined['日付'] >= start_a) & (df_combined['日付'] <= end_a)].copy()
            
            days_before = (end_before - start_before).days + 1
            days_after = (end_after - start_after).days + 1

            # データ件数のチェック
            expected_readings_b = 24 * days_before
            actual_readings_b = df_before.shape[0]
            if df_before.empty or actual_readings_b < expected_readings_b * 0.95:
                 st.warning(f"⚠️ **施工前期間 ({start_b}～{end_b}) のデータ注意:** 期待されるデータ件数 {expected_readings_b} 件に対し、実際は {actual_readings_b} 件です。")
            
            expected_readings_a = 24 * days_after
            actual_readings_a = df_after.shape[0]
            if df_after.empty or actual_readings_a < expected_readings_a * 0.95:
                 st.warning(f"⚠️ **施工後期間 ({start_a}～{end_a}) のデータ注意:** 期待されるデータ件数 {expected_readings_a} 件に対し、実際は {actual_readings_a} 件です。")
                
            # --- d) Excel書き込み ---
            success = write_excel_reports(temp_excel_path, df_before, df_after, start_b, end_b, start_a, end_a, operating_hours, store_name)
            
            if not success:
                return 

            # --- e) ファイル保存とダウンロード ---
            today_date_str = datetime.date.today().strftime('%Y%m%d')
            new_file_name = f"{store_name}：電力報告書{today_date_str}.xlsx"
            
            final_path = os.path.join(temp_dir, new_file_name)
            os.rename(temp_excel_path, final_path)
            
            with open(final_path, "rb") as file:
                st.success("✅ 処理が完了しました！以下のボタンから報告書をダウンロードしてください。")
                st.download_button(
                    label="⬇️ 報告書ファイルをダウンロード",
                    data=file,
                    file_name=new_file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
        except Exception as e:
            st.error("🚨 実行中にエラーが発生しました。ファイル形式と入力値を確認してください。")
            st.exception(e)

if __name__ == "__main__":
    main_streamlit_app()
